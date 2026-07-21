#!/usr/bin/env python3
"""Generate a simple Chinese annotation package for Plan B2.

Annotators only see Chinese column names and Chinese dropdown choices.
The private id map keeps faction / champion / confidence hidden.
"""

from __future__ import annotations

import csv
import random
import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "data/processed/kimi_agent_inventory.csv"
OUT = ROOT / "artifacts/annotation/codability-v0.2-cn"
PRIVATE = OUT / "private"
SEED = 20260611

FIELDS = [
    "编号",
    "理由文本",
    "1. 里面有没有可以查证的事实？",
    "2. 主要理由能不能被事实检查？",
    "3. 它主要属于哪类理由？",
    "4. 如果要检查，大概要查什么？",
    "5. 这条理由最适合放哪里？",
    "6. 你对自己的判断有多确定？",
    "主要理由摘录（可选）",
    "备注（可选）",
]

CHOICES = {
    "1. 里面有没有可以查证的事实？": [
        "有，明确可查",
        "有一点，但需要转成指标",
        "没有，主要是观点或感觉",
        "不确定",
    ],
    "2. 主要理由能不能被事实检查？": [
        "能，比较明确",
        "勉强能，需要补充规则",
        "很难，只能当叙事",
        "不能",
        "不确定",
    ],
    "3. 它主要属于哪类理由？": [
        "球员、阵容、年龄、身价",
        "排名、历史战绩、近期成绩",
        "伤病、赛程、体能、旅行",
        "赔率或市场观点",
        "战术、教练、打法",
        "心理、团队氛围、抗压",
        "模型自己说的概率",
        "传闻或朋友说",
        "玄学、周期、命理",
        "空泛判断",
        "其他或不确定",
    ],
    "4. 如果要检查，大概要查什么？": [
        "官方数据、排名、赛程",
        "比赛技术统计",
        "新闻或媒体报道",
        "赔率或市场页面",
        "模型原文即可",
        "找不到合适来源",
        "不确定",
    ],
    "5. 这条理由最适合放哪里？": [
        "可以进因子账本",
        "先当候选，之后再清洗",
        "只适合做边注",
        "没什么价值，可以丢弃",
        "不确定",
    ],
    "6. 你对自己的判断有多确定？": [
        "高",
        "中",
        "低",
    ],
}

HELP = [
    ["任务", "判断每条 AI 理由是否可以被事实检查。"],
    ["最重要", "不要查资料；不要判断真假；只看这句话是否原则上能被检查。"],
    ["例子1", "“法国队身价14.8亿欧”是可以查的事实。"],
    ["例子2", "“球队气氛空前团结”很难检查，通常只适合当叙事。"],
    ["例子3", "“某模型输出21.4%”只能说明模型这么说过，不等于事实。"],
    ["填写方式", "只填写黄色列。优先用下拉选项；备注可以不写。"],
]

EXAMPLES = [
    [
        "法国14.8亿欧身价，模型输出21.4%最高。",
        "第1题选“有，明确可查”；第2题多半选“勉强能，需要补充规则”；第5题选“先当候选，之后再清洗”。",
    ],
    [
        "南美球队热带基因觉醒，冠军气场回归。",
        "第1题选“没有，主要是观点或感觉”；第2题选“不能”；第5题选“只适合做边注”或“没什么价值”。",
    ],
    [
        "C罗41岁第六次出征，葡萄牙队内气氛空前团结。",
        "年龄和出征次数可查，但“气氛空前团结”难查；第5题通常选“只适合做边注”。",
    ],
]


def read_rows() -> list[dict[str, str]]:
    with INPUT.open(newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    if len(rows) != 300:
        raise ValueError(f"Expected 300 rows, got {len(rows)}")

    shuffled = rows[:]
    random.Random(SEED).shuffle(shuffled)
    for idx, row in enumerate(shuffled, start=1):
        row["编号"] = f"R{idx:04d}"
    return shuffled


def group_by_faction(rows: list[dict[str, str]]) -> dict[str, list[dict[str, str]]]:
    grouped: dict[str, list[dict[str, str]]] = {}
    for row in rows:
        grouped.setdefault(row["faction"], []).append(row)
    return grouped


def choose_samples(rows: list[dict[str, str]]) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    rng = random.Random(SEED + 1)
    calibration: list[dict[str, str]] = []
    kappa: list[dict[str, str]] = []
    for faction, items in sorted(group_by_faction(rows).items()):
        items = items[:]
        rng.shuffle(items)
        calibration.extend(items[:2])
        kappa.extend(items[2:12])
    rng.shuffle(calibration)
    rng.shuffle(kappa)
    return calibration, kappa


def add_info_sheet(wb: Workbook, title: str, row_count: int) -> None:
    ws = wb.create_sheet("先看这里", 0)
    ws.append([title])
    ws.append([f"本文件一共有 {row_count} 条。请到“标注表”工作表填写黄色列。"])
    ws.append([])
    for row in HELP:
        ws.append(row)
    ws.append([])
    ws.append(["字段怎么理解", "说明"])
    for field in FIELDS[2:8]:
        ws.append([field, "从下拉选项里选一个最接近的答案。"])
    ws.append(["主要理由摘录（可选）", "如果你愿意，可以用一句话写出这条理由最核心的说法。"])
    ws.append(["备注（可选）", "只有不确定、难选、想解释时再写。"])
    ws.append([])
    ws.append(["例子", "建议"])
    for row in EXAMPLES:
        ws.append(row)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 92
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws["A1"].font = Font(bold=True, size=16)


def add_choices_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("选项")
    max_len = max(len(v) for v in CHOICES.values())
    for col_idx, (field, choices) in enumerate(CHOICES.items(), start=1):
        ws.cell(row=1, column=col_idx, value=field)
        ws.cell(row=1, column=col_idx).font = Font(bold=True)
        for row_idx in range(max_len):
            ws.cell(row=row_idx + 2, column=col_idx, value=choices[row_idx] if row_idx < len(choices) else "")
    ws.sheet_state = "hidden"


def create_workbook(path: Path, rows: list[dict[str, str]], title: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "标注表"
    ws.append(FIELDS)

    for row in rows:
        ws.append([row["编号"], row["reason"], "", "", "", "", "", "", "", ""])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    locked_fill = PatternFill("solid", fgColor="E7E6E6")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            cell.fill = input_fill if cell.column >= 3 else locked_fill

    widths = {
        "A": 10,
        "B": 74,
        "C": 26,
        "D": 28,
        "E": 28,
        "F": 28,
        "G": 26,
        "H": 24,
        "I": 42,
        "J": 44,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions

    add_info_sheet(wb, title, len(rows))
    add_choices_sheet(wb)

    choices_ws = wb["选项"]
    for col_idx, field in enumerate(CHOICES.keys(), start=1):
        target_col = FIELDS.index(field) + 1
        col_letter = choices_ws.cell(row=1, column=col_idx).column_letter
        formula = f"=选项!${col_letter}$2:${col_letter}${len(CHOICES[field]) + 1}"
        dv = DataValidation(type="list", formula1=formula, allow_blank=False)
        dv.error = "请从下拉选项中选择。"
        dv.errorTitle = "无效选项"
        ws.add_data_validation(dv)
        dv.add(f"{ws.cell(row=2, column=target_col).coordinate}:{ws.cell(row=len(rows) + 1, column=target_col).coordinate}")

    wb.save(path)


def write_csv(path: Path, rows: list[dict[str, str]], fields: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def write_text_files() -> None:
    (OUT / "README.md").write_text(
        """# 可审计性标注包 v0.2 中文简化版

这套文件用来判断 AI 给出的世界杯预测理由，是否可以被事实检查。

## 发给标注者的文件

- `先试标20条.xlsx`
- `正式标注A-300条.xlsx`
- `复核标注B-100条.xlsx`
- `标注说明-请先读.md`
- `发送指南.md`

## 不要发送

- `private/` 目录里的任何文件。

## 标注者要记住

不查资料，不判断真假，只判断这句话是否原则上能被检查。
""",
        encoding="utf-8",
    )

    (OUT / "标注说明-请先读.md").write_text(
        """# 标注说明：判断一句话能不能被事实检查

## 你要做什么

你会看到很多条 AI 写的世界杯预测理由。请判断：这句话里面的理由，能不能被事实检查。

## 三条规则

1. **不要查资料。**
2. **不要判断它是真是假。**
3. **只判断它是否原则上可以被检查。**

## 怎么理解“可以检查”

可以检查：

- 提到球员年龄、阵容、身价、排名。
- 提到历史战绩、近期成绩、预选赛结果。
- 提到伤病、赛程、体能、旅行。
- 提到赔率或市场价格。

很难检查：

- “冠军气质很强”
- “队内气氛空前团结”
- “某队有玄学周期”
- “民族血性不在乎排名”
- “朋友说他们状态很好”

## 最容易混淆的情况

一句话里可能有一部分能查，一部分不能查。

例如：

> C罗41岁第六次出征，葡萄牙队内气氛空前团结。

“41岁”“第六次出征”可以查；但“气氛空前团结”很难查。

这种情况下，请你判断：这句话真正想用来支持预测的主要理由是什么。

## 文件怎么填

- 只填写黄色列。
- 大多数列直接从下拉选项里选。
- “主要理由摘录”和“备注”可以不写。
- 不确定时可以选“不确定”。
""",
        encoding="utf-8",
    )

    (OUT / "发送指南.md").write_text(
        """# 发送指南

## 第一步：先发试标

先发给对方：

- `先试标20条.xlsx`
- `标注说明-请先读.md`

发送话术：

```text
你好，我在做一个世界杯 AI 预测理由的小研究，需要你帮忙试标 20 条很短的中文句子。

你不需要查资料，也不需要判断句子是真是假。
只需要判断：这句话里的理由，原则上能不能被事实检查。

请打开 Excel，先看“先看这里”这个工作表，然后到“标注表”里填写黄色列。
大部分都是下拉选择，备注可以不写。

标完后把 Excel 原文件发回我就行。
```

## 第二步：确认理解没问题后，再发正式文件

主标注者发：

- `正式标注A-300条.xlsx`
- `标注说明-请先读.md`

复核标注者发：

- `复核标注B-100条.xlsx`
- `标注说明-请先读.md`

## 不要发

- 不要发 `private/` 目录。
- 不要发任何带“私有映射”的文件。

## 回收要求

让对方直接发回 Excel 原文件，不要截图，不要导出 PDF，不要改成 Word。
""",
        encoding="utf-8",
    )

    (OUT / "问卷字段备选.md").write_text(
        """# 如果改成问卷，字段这样设

推荐还是用 Excel。若必须用问卷，每条理由都问下面 6 个选择题。

1. 里面有没有可以查证的事实？
2. 主要理由能不能被事实检查？
3. 它主要属于哪类理由？
4. 如果要检查，大概要查什么？
5. 这条理由最适合放哪里？
6. 你对自己的判断有多确定？

另外加两个可选填空：

- 主要理由摘录
- 备注
""",
        encoding="utf-8",
    )

    (OUT / "快速说明.html").write_text(
        """<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <title>标注快速说明</title>
  <style>
    body { font-family: -apple-system, BlinkMacSystemFont, "PingFang SC", "Microsoft YaHei", sans-serif; max-width: 880px; margin: 40px auto; line-height: 1.75; color: #172033; }
    h1, h2 { color: #163b5c; }
    .box { background: #fff7e6; border-left: 4px solid #d88900; padding: 12px 16px; margin: 18px 0; }
    li { margin: 8px 0; }
  </style>
</head>
<body>
  <h1>标注快速说明</h1>
  <div class="box">不要查资料；不要判断真假；只判断这句话是否原则上能被事实检查。</div>
  <h2>可以检查的例子</h2>
  <ul>
    <li>球员年龄、阵容、身价、排名</li>
    <li>历史战绩、近期成绩、预选赛结果</li>
    <li>伤病、赛程、体能、旅行</li>
    <li>赔率或市场价格</li>
  </ul>
  <h2>很难检查的例子</h2>
  <ul>
    <li>冠军气质、心理玄学、队内氛围</li>
    <li>周期轮回、命理、朋友说、空泛判断</li>
  </ul>
</body>
</html>
""",
        encoding="utf-8",
    )

    codebook_rows = []
    for field, choices in CHOICES.items():
        for choice in choices:
            codebook_rows.append({"题目": field, "选项": choice})
    write_csv(OUT / "标注选项表.csv", codebook_rows, ["题目", "选项"])


def main() -> int:
    if OUT.exists():
        shutil.rmtree(OUT)
    OUT.mkdir(parents=True)
    PRIVATE.mkdir(parents=True)

    rows = read_rows()
    calibration, kappa = choose_samples(rows)

    write_csv(
        PRIVATE / "私有映射-不要发送.csv",
        rows,
        ["编号", "agent_id", "faction", "persona", "champion", "top3", "confidence", "source_file", "reason"],
    )

    create_workbook(OUT / "先试标20条.xlsx", calibration, "先试标 20 条")
    create_workbook(OUT / "正式标注A-300条.xlsx", rows, "正式标注 A：300 条")
    create_workbook(OUT / "复核标注B-100条.xlsx", kappa, "复核标注 B：100 条")
    write_text_files()

    print(f"生成完成：{OUT}")
    print(f"行数：试标={len(calibration)}，正式={len(rows)}，复核={len(kappa)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
